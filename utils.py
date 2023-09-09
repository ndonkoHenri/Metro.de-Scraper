import os
import openpyxl
from splinter import Browser
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Color
import flet as ft


class MetroScraper:

    def __init__(
            self,
            browser: Browser,
            page: ft.Page = None,
            source_path: str = "source.xlsx",
            destination_path: str = "Result.xlsx"
    ):
        self.browser: Browser = None
        self.failures = []
        self.error_counter = 0
        self.logs_cache = []

        self.work_sheet = None
        self.work_book = None

        self.name_xpath = "//div[@class='mfcss_article-detail--title']//child::span"
        self.id_xpath = "(//div[@class='mfcss_article-detail--articlenumber']//descendant::span)[1]"
        self.price_xpath = "(//span[@class='mfcss_article-detail--price-breakdown primary']//child::span)[2]"
        self.price_if_promo_xpath = "//span[@class='mfcss_article-detail--price-breakdown strike']//span//span"
        # price_2_xpath = "(//span[@class='mfcss_article-detail--price-breakdown primary promotion']//child::span)[2]"

        self.start_automation(browser, page, source_path, destination_path)

    def get_item_info(self, iid, link):
        """
        Uses the xpaths provided to find the name of the product, its price (if it is on sale),
        and whether it is on sale. It returns these values as well as its own ID.

        Args:
            iid: the id from the excel file
            link: direct link to the article's page

        :returns A tuple with the item's id, name, price and a boolean that indicates if it is on promotion/sale
        """
        item_id, item_name, item_price, is_promotion = iid, None, None, False

        try:
            item_name = self.browser.find_by_xpath(self.name_xpath).first.value
            item_id = self.browser.find_by_xpath(self.id_xpath).first.value

            # if no price is found using price_xpath, then the article is certainly on promotion, hence we try price_if_promo_xpath
            try:
                item_price = self.browser.find_by_xpath(self.price_xpath).first.value
            except Exception as e:
                item_price = self.browser.find_by_xpath(self.price_if_promo_xpath).first.value
                is_promotion = True

            # the id on the webpage and that from the source excel file are different
            if item_id != str(iid):
                # write_log(f"! ids are different: Online={item_id} - Local={iid=} - {link=}")
                pass

        except Exception as e:
            self.write_log(f"Error: {iid} ~ {link} ~ {e} ")

        return item_id, item_name, item_price, is_promotion

    def read_source(self, src_path):
        """
        Reads the source Excel file of this project and stores the data in a pandas dataframe.
        The first column of the dataframe with header row 'source' contains the needed data.

        :param src_path: path to the source file
        """
        df = pd.read_excel(src_path)

        article_number_col = df[
            "Metro Artikelnummer"]  # important: make sure the name of the first column is correctly written
        links_col = df["Link"]  # important: make sure the name of the second column is correctly written

        # the number of rows in the excel file - this will be used to show operation progression
        max_row_count = max(df.count().to_list())

        return article_number_col, links_col, max_row_count

    def summarize(self, td):
        """
        Reports the total execution time, number of errors, and failures in to the log file.

        :param td: Pass in the total execution time of the program
        """
        self.write_log(f'Total execution time: {td} seconds or {round(td / 60, 1)} minutes')
        self.write_log(f'Number of errors: {self.error_counter}')
        # self.write_log(f"Failures:\n{self.failures}")

    def reject_cookies(self):
        """Clicks on the button in the cookie banner to reject cookies. The banner is found in the shadow DOM."""
        time.sleep(3)

        try:
            shadow = self.browser.find_by_css("cms-cookie-disclaimer").first.shadow_root
        except:
            self.write_log("Cookie Refusal failed. Retrying..")
            time.sleep(5)
            shadow = self.browser.find_by_css("cms-cookie-disclaimer").first.shadow_root

        shadow.find_by_css(".btn-secondary.reject-btn.field-reject-button-name").click()

    def prepare_result_worksheet(self, dst_path):
        """
        Prepares the result's workbook.
        If the file already exists, it will be opened/used and a new sheet will be created.
        Otherwise, a new Workbook is created with an active Worksheet.

        :param dst_path: the path to the result's excel file
        """
        sheet_name = f"{time.strftime('%d.%m.%Y - %Hh%M')}"

        if os.path.exists(dst_path):
            self.work_book = openpyxl.load_workbook(dst_path)
            self.work_sheet = self.work_book.create_sheet(sheet_name)
            self.work_book.active = self.work_sheet
        else:
            self.work_book = Workbook()
            self.work_sheet = self.work_book.active
            self.work_sheet.title = sheet_name

        # add header row
        self.work_sheet.append(['ID', 'Name', 'Preis'])

        # customize header row
        font = Font(bold=True, size=11)
        for r in self.work_sheet["A1:C1"]:
            for s in r:
                s.font = font

    def save_results(self, dst_path):
        """
        Saves the results of the scraping to an excel file.

        :param dst_path: the path where the results should be saved
        """
        try:
            self.work_book.save(dst_path)
        except PermissionError:
            new_path = dst_path.replace(dst_path.split("\\")[-1], f"Result {time.strftime('%Hh%M')}.xlsx")
            self.work_book.save(new_path)
            self.write_log(
                f"{'Permission errors were encountered while trying to save the results at the chosen destination. Please, always make sure to close the excel files in use.'.upper()} Result file was instead saved at {new_path}")

    def write_log(self, text):
        """Takes a string as an argument and writes it to the logs.txt file."""
        with open("logs.txt", "a", encoding="utf-8") as f:
            f.write(f"\n- {text}")

        self.logs_cache.append(f"- {text}\n")

    def start_automation(
            self,
            browser: Browser,
            page: ft.Page = None,
            source_path: str = "source.xlsx",
            destination_path: str = "Result.xlsx"
    ):
        """
        Main start point.
        It calls the necessary functions, starts the scraping, saves and summarizes.
        """
        start_time = time.time()
        articles_in_promotion = []

        self.browser = browser

        try:
            self.write_log(f"Automation Initialization on the {time.strftime('%d.%m.%Y - %H:%M %p')}")
            self.write_log(f"{source_path=} | {destination_path=}")

            article_number_col, links_col, max_row_count = self.read_source(source_path)

            self.browser.visit("https://produkte.metro.de/shop/")

            self.reject_cookies()

            self.prepare_result_worksheet(destination_path)

            # set the progress bar value to zero (operation initialisation)
            if page:
                page.splash.value = page.window_progress_bar = 0
                page.update()

            for count, source_info in enumerate(zip(article_number_col, links_col), start=2):
                s_number, s_link = source_info
                if s_link.startswith("https://produkte.metro.de/shop/"):
                    self.browser.visit(s_link)

                    iid, name, price, is_promotion = self.get_item_info(s_number, s_link)

                    # if name or price is None (was not scraped successfully), retry several times before giving up
                    loop = 0
                    while ((name is None) or (price is None)) and (loop != 10):
                        time.sleep(5)
                        iid, name, price, is_promotion = self.get_item_info(s_number, s_link)

                        # refresh the web page at certain intervals (useful when the page didn't initially load completely)
                        if (loop in [1, 4, 7]) and ((name is None) or (price is None)):
                            self.browser.reload()

                        loop += 1

                        time.sleep(2.5)

                    # while loop was called, and the error was then resolved
                    if (name is not None) and (price is not None) and loop > 0:
                        self.write_log("RESOLVED error above!")
                        self.work_sheet.append([iid, str(name), str(price)])

                    # while loop was called, and the error was NOT resolved
                    elif ((name is None) or (price is None)) and loop > 0:
                        self.error_counter += 1
                        self.write_log(f"FAILED after while loop: {iid, s_link}")
                        self.failures.append((iid, s_link))
                        self.work_sheet.append([iid, str(name), str(price), s_link])

                    # went smoothly with no errors - while loop not called (loop==0 is True)
                    else:
                        self.work_sheet.append([iid, str(name), str(price)])

                    if is_promotion:
                        articles_in_promotion.append(count)

                # if the link is not a metro link, ignore
                else:
                    self.work_sheet.append([s_number, "IGNORED", "IGNORED"])

                # update the progress bar value to reflect progression
                if page:
                    page.splash.value = page.window_progress_bar = (count - 1) / max_row_count
                    page.update()

            # set the progress bar value to 1 (operation completed)
            if page:
                page.splash.value = 1
                page.update()

            # if there are some articles in sales/promotion -> change the color of the price to red (kind of notifier)
            if articles_in_promotion:
                for i in articles_in_promotion:
                    cell = self.work_sheet.cell(row=i, column=3)
                    cell.font = Font(color=Color(rgb="FF0000"))

            self.save_results(destination_path)

            time_delta = round(time.time() - start_time, None)  # the total duration of the execution/automation
            self.summarize(time_delta)

        except Exception as e:
            self.write_log(e)
            self.write_log(f"{time.strftime('%H:%M %p')} | Seems like something went wrong :(, please retry.")
        finally:
            self.browser.quit()  # close browser session
            return self.logs_cache


if __name__ == "__main__":
    MetroScraper(Browser("chrome"))
